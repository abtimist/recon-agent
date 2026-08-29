"""
core/report_generator.py

Pure-Python report generation for Recon Agent reconciliation results.

Accepts plain Python dicts (serialized ReconcileResult / BatchReconcileResult)
and returns io.BytesIO blobs.  No FastAPI dependency, no DB calls, no
reconciliation logic — this module is exclusively a formatter.

Supports:
  - generate_single_excel(result)   → bytes
  - generate_single_pdf(result)     → bytes
  - generate_batch_excel(batch)     → bytes
  - generate_batch_pdf(batch)       → bytes
"""

from __future__ import annotations

import io
import copy
from datetime import datetime, timezone
from typing import Any

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _fmt_inr(value: float | None) -> str:
    """Format a number as ₹1,23,456.78 (Indian grouping style)."""
    if value is None:
        return "—"
    try:
        v = float(value)
        # Python locale-independent Indian grouping
        s = f"{abs(v):,.2f}"
        return f"₹{s}" if v >= 0 else f"−₹{s}"
    except (TypeError, ValueError):
        return "—"

def _fmt_pct(value: float | None) -> str:
    if value is None:
        return "—"
    return f"{value:.2f}%"

def _fmt_date(value: str | None) -> str:
    if not value:
        return ""
    try:
        dt = datetime.fromisoformat(str(value).replace("Z", "+00:00"))
        return dt.strftime("%d %b %Y, %H:%M UTC")
    except Exception:
        return str(value).split("T")[0]

def _now_str() -> str:
    return datetime.now(timezone.utc).strftime("%d %b %Y, %H:%M UTC")

def _exc_type_label(exc_type: str) -> str:
    labels = {
        "missing_target_record":     "Missing in Target",
        "stray_target_record":       "Extra in Target",
        "ai_rejected_ambiguous_match": "AI Rejected",
    }
    return labels.get(exc_type, exc_type)


# ===========================================================================
# EXCEL
# ===========================================================================

def _excel_styles():
    """Return a lazy-loaded style dict so openpyxl imports stay inside funcs."""
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter

    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    return {
        "title":   Font(name="Calibri", bold=True, size=14, color="1A1A1A"),
        "header":  Font(name="Calibri", bold=True, size=10, color="FFFFFF"),
        "body":    Font(name="Calibri", size=10, color="1A1A1A"),
        "muted":   Font(name="Calibri", size=9, color="888888"),
        "accent":  Font(name="Calibri", bold=True, size=10, color="2D7D46"),
        "danger":  Font(name="Calibri", bold=True, size=10, color="C0392B"),
        "orange":  Font(name="Calibri", bold=True, size=10, color="D35400"),

        "fill_header":   PatternFill("solid", fgColor="1A1A1A"),
        "fill_summary":  PatternFill("solid", fgColor="F4F6F8"),
        "fill_alt":      PatternFill("solid", fgColor="FAFBFC"),
        "fill_red":      PatternFill("solid", fgColor="FDEDEC"),
        "fill_blue":     PatternFill("solid", fgColor="EBF5FB"),
        "fill_amber":    PatternFill("solid", fgColor="FDEBD0"),
        "fill_green":    PatternFill("solid", fgColor="EAFAF1"),
        "fill_title":    PatternFill("solid", fgColor="F0FFF4"),

        "center": Alignment(horizontal="center", vertical="center"),
        "left":   Alignment(horizontal="left",   vertical="center"),
        "right":  Alignment(horizontal="right",  vertical="center"),
        "wrap":   Alignment(wrap_text=True, vertical="top"),
        "border": border,

        "fmt_inr": '#,##0.00',
        "fmt_pct": '0.00"%"',
        "fmt_int": '#,##0',
    }


def _apply_header_row(ws, row_idx: int, columns: list[str], st: dict) -> None:
    from openpyxl.styles import PatternFill
    for col_idx, label in enumerate(columns, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=label)
        cell.font      = st["header"]
        cell.fill      = st["fill_header"]
        cell.alignment = st["center"]
        cell.border    = st["border"]


def _autofit_columns(ws, min_w: int = 10, max_w: int = 45) -> None:
    from openpyxl.utils import get_column_letter
    for col in ws.columns:
        best = min_w
        for cell in col:
            try:
                if cell.value:
                    best = max(best, min(len(str(cell.value)) + 2, max_w))
            except Exception:
                pass
        ws.column_dimensions[get_column_letter(col[0].column)].width = best


# ---------------------------------------------------------------------------
# Single — Summary sheet
# ---------------------------------------------------------------------------

def _write_single_summary_sheet(ws, result: dict, st: dict) -> None:
    ws.title = "Summary"

    # Branding / title row
    ws.merge_cells("A1:D1")
    title_cell = ws["A1"]
    title_cell.value     = "Recon Agent — Reconciliation Report"
    title_cell.font      = st["title"]
    title_cell.fill      = st["fill_title"]
    title_cell.alignment = st["center"]
    ws.row_dimensions[1].height = 28

    def row(label: str, value, bold_val: bool = False, fmt=None):
        r = ws.max_row + 1
        lc = ws.cell(row=r, column=1, value=label)
        vc = ws.cell(row=r, column=2, value=value)
        lc.font = st["muted"]
        vc.font = Font(name="Calibri", size=10, bold=bold_val, color="1A1A1A") if bold_val else st["body"]
        lc.alignment = st["left"]
        vc.alignment = st["left"]
        if fmt:
            vc.number_format = fmt
        return vc

    from openpyxl.styles import Font

    ws.append([])  # blank
    ws.append(["Run ID",          result.get("run_id", "—")])
    ws.append(["Source File",     result.get("source_filename", "—")])
    ws.append(["Target File",     result.get("target_filename", "—")])
    ws.append(["Reconciled At",   _fmt_date(result.get("completed_at"))])
    ws.append(["Report Generated", _now_str()])
    ws.append([])

    ws.append(["— MATCHING RULES —", ""])
    ws.append(["Amount Tolerance", result.get("amount_tolerance", 0)])
    ws.cell(ws.max_row, 2).number_format = st["fmt_inr"]
    ws.append(["Date Window (days)", result.get("date_window_days", 0)])
    ws.append(["AI Provider",    result.get("ai_provider", "none")])
    ws.append([])

    ws.append(["— MATCH RESULTS —", ""])
    ws.append(["Total Transactions",  result.get("total_source_rows", 0)])
    ws.cell(ws.max_row, 2).number_format = st["fmt_int"]
    ws.append(["Total Matched",       result.get("total_matched", 0)])
    ws.cell(ws.max_row, 2).number_format = st["fmt_int"]
    ws.append(["Match Rate",          (result.get("match_rate", 0) or 0) / 100])
    ws.cell(ws.max_row, 2).number_format = "0.00%"
    ws.append(["  · Exact Matches",   result.get("exact_matches", 0)])
    ws.cell(ws.max_row, 2).number_format = st["fmt_int"]
    ws.append(["  · Fuzzy Matches",   result.get("fuzzy_matches", 0)])
    ws.cell(ws.max_row, 2).number_format = st["fmt_int"]
    ws.append(["  · AI-Resolved",     result.get("ai_matches", 0)])
    ws.cell(ws.max_row, 2).number_format = st["fmt_int"]
    ws.append(["Total Exceptions",    result.get("exceptions_count", 0)])
    ws.cell(ws.max_row, 2).number_format = st["fmt_int"]
    ws.append([])

    summary = result.get("summary") or {}
    ws.append(["— FINANCIALS —", ""])
    ws.append(["Total Amount",     summary.get("total_amount", 0)])
    ws.cell(ws.max_row, 2).number_format = st["fmt_inr"]
    ws.append(["Matched Amount",   summary.get("matched_amount", 0)])
    ws.cell(ws.max_row, 2).number_format = st["fmt_inr"]
    ws.append(["Unmatched Amount", summary.get("unmatched_amount", 0)])
    ws.cell(ws.max_row, 2).number_format = st["fmt_inr"]
    ws.append([])

    # Style label column
    for r in ws.iter_rows(min_row=2):
        lc = r[0]
        if lc.value and str(lc.value).startswith("—"):
            lc.font      = Font(name="Calibri", bold=True, size=9, color="555555")
            lc.fill      = st["fill_summary"]
            if len(r) > 1:
                r[1].fill = st["fill_summary"]
        else:
            lc.font      = st["muted"]

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 32


# ---------------------------------------------------------------------------
# Single — Exceptions sheet
# ---------------------------------------------------------------------------

def _write_exceptions_sheet(ws, exceptions: list[dict], st: dict, sheet_title: str = "Exceptions") -> None:
    ws.title = sheet_title
    cols = ["#", "Type", "Transaction ID", "Party / Merchant", "Amount (₹)", "Date", "Reason"]
    _apply_header_row(ws, 1, cols, st)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    fill_map = {
        "missing_target_record":       st["fill_red"],
        "stray_target_record":         st["fill_blue"],
        "ai_rejected_ambiguous_match": st["fill_amber"],
    }

    for i, exc in enumerate(exceptions, start=1):
        exc_type = exc.get("type", "")
        fill = fill_map.get(exc_type, st["fill_alt"])
        row_data = [
            i,
            _exc_type_label(exc_type),
            str(exc.get("id", "")),
            str(exc.get("party", "")),
            float(exc.get("amount", 0) or 0),
            str(exc.get("date", "")).split("T")[0].split(" ")[0],
            str(exc.get("reason", "")),
        ]
        ws.append(row_data)
        r = ws.max_row
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=r, column=col_idx)
            cell.fill      = fill
            cell.alignment = st["left"]
            cell.border    = st["border"]
            cell.font      = st["body"]
        ws.cell(row=r, column=5).number_format = st["fmt_inr"]

    col_widths = [6, 22, 24, 28, 18, 14, 60]
    from openpyxl.utils import get_column_letter
    for ci, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w


# ---------------------------------------------------------------------------
# Single — Duplicates sheet
# ---------------------------------------------------------------------------

def _write_duplicates_sheet(ws, dup_report: dict, st: dict) -> None:
    ws.title = "Duplicates"
    if not dup_report:
        ws.append(["No duplicate detection data available."])
        return

    cols = ["File", "Party / Merchant", "Amount (₹)", "Date", "Occurrences", "Row IDs"]
    _apply_header_row(ws, 1, cols, st)
    ws.freeze_panes = "A2"

    def write_groups(groups: list[dict], label: str):
        for grp in groups:
            row_data = [
                label,
                str(grp.get("party", "")),
                float(grp.get("amount", 0) or 0),
                str(grp.get("date", "")),
                int(grp.get("occurrences", 0)),
                ", ".join(str(x) for x in (grp.get("row_ids") or [])),
            ]
            ws.append(row_data)
            r = ws.max_row
            for ci, _ in enumerate(row_data, start=1):
                c = ws.cell(row=r, column=ci)
                c.fill      = st["fill_amber"]
                c.alignment = st["left"]
                c.border    = st["border"]
                c.font      = st["body"]
            ws.cell(row=r, column=3).number_format = st["fmt_inr"]

    write_groups(dup_report.get("source") or [], "Source")
    write_groups(dup_report.get("target") or [], "Target")

    if ws.max_row == 1:
        ws.append(["No duplicate groups found.", "", "", "", "", ""])

    from openpyxl.utils import get_column_letter
    for ci, w in enumerate([10, 28, 18, 14, 14, 50], start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w


# ---------------------------------------------------------------------------
# Single — Top Merchants sheet
# ---------------------------------------------------------------------------

def _write_top_merchants_sheet(ws, summary: dict, st: dict) -> None:
    from openpyxl.styles import PatternFill
    ws.title = "Top Exception Merchants"
    merchants = summary.get("top_exception_merchants") or []

    cols = ["Rank", "Party / Merchant", "Exception Count"]
    _apply_header_row(ws, 1, cols, st)
    ws.freeze_panes = "A2"

    white_fill = PatternFill("solid", fgColor="FFFFFF")
    for i, m in enumerate(merchants, start=1):
        ws.append([i, str(m.get("party", "")), int(m.get("count", 0))])
        r = ws.max_row
        for ci in range(1, 4):
            c = ws.cell(row=r, column=ci)
            c.fill      = st["fill_alt"] if i % 2 == 0 else white_fill
            c.alignment = st["center"] if ci in (1, 3) else st["left"]
            c.border    = st["border"]
            c.font      = st["body"]

    from openpyxl.styles import PatternFill
    from openpyxl.utils import get_column_letter
    for ci, w in enumerate([8, 38, 18], start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w


# ---------------------------------------------------------------------------
# Public: generate_single_excel
# ---------------------------------------------------------------------------

def generate_single_excel(result: dict) -> bytes:
    """Generate a professional Excel workbook for a single reconciliation run."""
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill  # needed inside sub-fns
    result = copy.deepcopy(result)

    wb = Workbook()
    st = _excel_styles()

    ws_summary = wb.active
    _write_single_summary_sheet(ws_summary, result, st)

    ws_exc = wb.create_sheet()
    _write_exceptions_sheet(ws_exc, result.get("exception_report") or [], st)

    ws_dup = wb.create_sheet()
    _write_duplicates_sheet(ws_dup, result.get("duplicates") or {}, st)

    ws_merch = wb.create_sheet()
    _write_top_merchants_sheet(ws_merch, result.get("summary") or {}, st)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Public: generate_batch_excel
# ---------------------------------------------------------------------------

def generate_batch_excel(batch: dict) -> bytes:
    """Generate a multi-sheet Excel workbook for a batch reconciliation run."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    batch = copy.deepcopy(batch)
    st   = _excel_styles()
    wb   = Workbook()

    # -------- Sheet 1: Batch Overview --------
    ws_ov = wb.active
    ws_ov.title = "Batch Overview"
    summary = batch.get("summary") or {}

    ws_ov.merge_cells("A1:E1")
    ws_ov["A1"].value     = "Recon Agent — Batch Reconciliation Report"
    ws_ov["A1"].font      = st["title"]
    ws_ov["A1"].fill      = st["fill_title"]
    ws_ov["A1"].alignment = st["center"]
    ws_ov.row_dimensions[1].height = 28

    ws_ov.append([])
    ws_ov.append(["Report Generated", _now_str()])
    ws_ov.append(["File Pairs",       summary.get("completed_runs", 0) + summary.get("failed_runs", 0)])
    ws_ov.append(["Completed",        summary.get("completed_runs", 0)])
    ws_ov.append(["Failed",           summary.get("failed_runs", 0)])
    ws_ov.append([])
    ws_ov.append(["— OVERALL RESULTS —", ""])
    ws_ov.append(["Total Transactions",  summary.get("total_transactions", 0)])
    ws_ov.cell(ws_ov.max_row, 2).number_format = st["fmt_int"]
    ws_ov.append(["Total Matched",       summary.get("total_matched", 0)])
    ws_ov.cell(ws_ov.max_row, 2).number_format = st["fmt_int"]
    ws_ov.append(["Overall Match Rate",  (summary.get("overall_match_rate", 0) or 0) / 100])
    ws_ov.cell(ws_ov.max_row, 2).number_format = "0.00%"
    ws_ov.append(["Total Exceptions",    summary.get("total_exceptions", 0)])
    ws_ov.cell(ws_ov.max_row, 2).number_format = st["fmt_int"]
    ws_ov.append([])
    ws_ov.append(["— FINANCIALS —", ""])
    ws_ov.append(["Total Amount",     summary.get("total_amount", 0)])
    ws_ov.cell(ws_ov.max_row, 2).number_format = st["fmt_inr"]
    ws_ov.append(["Matched Amount",   summary.get("total_matched_amount", 0)])
    ws_ov.cell(ws_ov.max_row, 2).number_format = st["fmt_inr"]
    ws_ov.append(["Unmatched Amount", summary.get("total_unmatched_amount", 0)])
    ws_ov.cell(ws_ov.max_row, 2).number_format = st["fmt_inr"]
    ws_ov.append([])
    ws_ov.append(["Duplicate Groups (Source)", summary.get("duplicate_source_groups", 0)])
    ws_ov.append(["Duplicate Groups (Target)", summary.get("duplicate_target_groups", 0)])

    ws_ov.column_dimensions["A"].width = 30
    ws_ov.column_dimensions["B"].width = 35

    # -------- Sheet 2: Run Summary --------
    ws_runs = wb.create_sheet("Run Summary")
    run_cols = [
        "#", "Source File", "Target File", "Status",
        "Transactions", "Matched", "Match Rate",
        "Exceptions", "Total Amount (₹)", "Matched Amount (₹)", "Unmatched Amount (₹)",
        "Dup Groups (Src)", "Dup Groups (Tgt)", "Error"
    ]
    _apply_header_row(ws_runs, 1, run_cols, st)
    ws_runs.freeze_panes = "A2"
    ws_runs.auto_filter.ref = "A1:N1"

    runs = batch.get("runs") or []
    for i, run in enumerate(runs, start=1):
        status     = run.get("status", "failed")
        r_data     = run.get("result") or {}
        run_sum    = r_data.get("summary") or {}
        run_dups   = r_data.get("duplicates") or {}

        if status == "completed":
            row_vals = [
                i,
                run.get("source_filename", ""),
                run.get("target_filename", ""),
                "COMPLETED",
                r_data.get("total_source_rows", 0),
                r_data.get("total_matched", 0),
                (r_data.get("match_rate", 0) or 0) / 100,
                r_data.get("exceptions_count", 0),
                run_sum.get("total_amount", 0),
                run_sum.get("matched_amount", 0),
                run_sum.get("unmatched_amount", 0),
                run_dups.get("source_count", 0),
                run_dups.get("target_count", 0),
                "",
            ]
            fill = st["fill_alt"] if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        else:
            row_vals = [
                i,
                run.get("source_filename", ""),
                run.get("target_filename", ""),
                "FAILED",
                "—", "—", "—", "—", "—", "—", "—", "—", "—",
                str(run.get("error", "")),
            ]
            fill = st["fill_red"]

        ws_runs.append(row_vals)
        r = ws_runs.max_row
        for ci, _ in enumerate(row_vals, start=1):
            c = ws_runs.cell(row=r, column=ci)
            c.fill      = fill
            c.alignment = st["left"]
            c.border    = st["border"]
            c.font      = st["body"]

        if status == "completed":
            ws_runs.cell(r, 7).number_format  = "0.00%"
            ws_runs.cell(r, 9).number_format  = st["fmt_inr"]
            ws_runs.cell(r, 10).number_format = st["fmt_inr"]
            ws_runs.cell(r, 11).number_format = st["fmt_inr"]

    for ci, w in enumerate([5, 28, 28, 12, 14, 12, 14, 12, 20, 20, 20, 16, 16, 50], start=1):
        ws_runs.column_dimensions[get_column_letter(ci)].width = w

    # -------- Sheets 3+: Per-run exception details --------
    for i, run in enumerate(runs, start=1):
        if run.get("status") != "completed":
            continue
        r_data     = run.get("result") or {}
        exceptions = r_data.get("exception_report") or []
        if not exceptions:
            continue

        src_name = run.get("source_filename", f"run_{i}")
        # Excel sheet names max 31 chars
        sheet_label = f"R{i:02d} – {src_name}"[:31]
        ws_det = wb.create_sheet(sheet_label)
        _write_exceptions_sheet(ws_det, exceptions, st, sheet_title=sheet_label)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ===========================================================================
# PDF
# ===========================================================================

def _pdf_styles():
    """Return ReportLab style sheet."""
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums  import TA_CENTER, TA_LEFT, TA_RIGHT
    from reportlab.lib        import colors

    ss = getSampleStyleSheet()

    DARK   = colors.HexColor("#1A1A1A")
    GREEN  = colors.HexColor("#2D7D46")
    RED    = colors.HexColor("#C0392B")
    AMBER  = colors.HexColor("#D35400")
    BLUE   = colors.HexColor("#1A5276")
    GREY   = colors.HexColor("#888888")
    LGREY  = colors.HexColor("#F4F6F8")
    ACCENT = colors.HexColor("#b3ff00")

    styles = {
        "title": ParagraphStyle(
            "ReportTitle", parent=ss["Heading1"],
            fontSize=20, textColor=DARK, spaceAfter=4, spaceBefore=0,
        ),
        "subtitle": ParagraphStyle(
            "ReportSubtitle", parent=ss["Normal"],
            fontSize=10, textColor=GREY, spaceAfter=12,
        ),
        "section": ParagraphStyle(
            "SectionHead", parent=ss["Normal"],
            fontSize=11, fontName="Helvetica-Bold", textColor=DARK,
            spaceBefore=16, spaceAfter=6,
            borderPad=4,
        ),
        "body": ParagraphStyle(
            "Body", parent=ss["Normal"],
            fontSize=9, textColor=DARK, spaceAfter=3,
        ),
        "muted": ParagraphStyle(
            "Muted", parent=ss["Normal"],
            fontSize=8, textColor=GREY, spaceAfter=2,
        ),
        "table_header": ParagraphStyle(
            "TH", parent=ss["Normal"],
            fontSize=8, fontName="Helvetica-Bold", textColor=colors.white,
        ),
        "table_cell": ParagraphStyle(
            "TC", parent=ss["Normal"],
            fontSize=8, textColor=DARK,
        ),
        "colors": {
            "dark": DARK, "green": GREEN, "red": RED,
            "amber": AMBER, "blue": BLUE, "grey": GREY,
            "lgrey": LGREY, "accent": ACCENT,
            "white": colors.white,
        }
    }
    return styles


def _pdf_kpi_table(data_pairs: list[tuple[str, str]], col_count: int = 3):
    """Build a KPI grid table from (label, value) pairs."""
    from reportlab.platypus import Table, TableStyle
    from reportlab.lib      import colors

    c = col_count
    # Pad to multiple of c
    while len(data_pairs) % c != 0:
        data_pairs.append(("", ""))

    rows = []
    for i in range(0, len(data_pairs), c):
        label_row  = [p[0] for p in data_pairs[i:i+c]]
        value_row  = [p[1] for p in data_pairs[i:i+c]]
        rows.append(label_row)
        rows.append(value_row)

    col_w = 160 / c
    tbl = Table(rows, colWidths=[col_w * 72 / 72] * c)  # points
    # Use fixed approximate col widths
    col_pts = [510 // c] * c

    tbl = Table(rows, colWidths=col_pts)
    style = TableStyle([
        ("BACKGROUND",   (0, 0), (-1, -1), colors.HexColor("#F4F6F8")),
        ("TEXTCOLOR",    (0, 0), (-1, 0),  colors.HexColor("#888888")),
        ("FONTSIZE",     (0, 0), (-1, 0),  8),
        ("FONTNAME",     (0, 1), (-1, -1), "Helvetica-Bold"),
        ("FONTSIZE",     (0, 1), (-1, -1), 14),
        ("TEXTCOLOR",    (0, 1), (-1, -1), colors.HexColor("#1A1A1A")),
        ("ALIGN",        (0, 0), (-1, -1), "CENTER"),
        ("VALIGN",       (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 0), (-1, -1), [colors.HexColor("#F4F6F8"), colors.white]),
        ("BOX",          (0, 0), (-1, -1), 0.5, colors.HexColor("#DDDDDD")),
        ("INNERGRID",    (0, 0), (-1, -1), 0.5, colors.HexColor("#DDDDDD")),
        ("TOPPADDING",   (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 8),
    ])
    tbl.setStyle(style)
    return tbl


def _pdf_exc_table(exceptions: list[dict], st: dict):
    from reportlab.platypus import Table, TableStyle, Paragraph
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib      import colors

    cl = st["colors"]
    header = ["#", "Type", "Transaction ID", "Merchant", "Amount", "Date", "Reason"]
    col_widths = [20, 80, 80, 80, 55, 50, 145]  # points, total ~510

    data = [header]
    fill_map = {
        "missing_target_record":       colors.HexColor("#FDEDEC"),
        "stray_target_record":         colors.HexColor("#EBF5FB"),
        "ai_rejected_ambiguous_match": colors.HexColor("#FDEBD0"),
    }
    row_fills = [cl["dark"]]  # header fill handled by style

    for i, exc in enumerate(exceptions[:200], start=1):   # cap at 200 rows in PDF
        exc_type = exc.get("type", "")
        row_fills.append(fill_map.get(exc_type, colors.HexColor("#FFFFFF")))
        data.append([
            str(i),
            _exc_type_label(exc_type),
            str(exc.get("id", ""))[:20],
            str(exc.get("party", ""))[:22],
            _fmt_inr(exc.get("amount")),
            str(exc.get("date", "")).split("T")[0].split(" ")[0],
            Paragraph(str(exc.get("reason", ""))[:120],
                      ParagraphStyle("er", fontSize=7, leading=9)),
        ])

    from reportlab.lib.styles import ParagraphStyle
    tbl = Table(data, colWidths=col_widths, repeatRows=1)
    style_cmds = [
        ("BACKGROUND",   (0, 0), (-1, 0),  colors.HexColor("#1A1A1A")),
        ("TEXTCOLOR",    (0, 0), (-1, 0),  colors.white),
        ("FONTNAME",     (0, 0), (-1, 0),  "Helvetica-Bold"),
        ("FONTSIZE",     (0, 0), (-1, 0),  8),
        ("FONTSIZE",     (0, 1), (-1, -1), 7),
        ("ALIGN",        (0, 0), (-1, -1), "LEFT"),
        ("VALIGN",       (0, 0), (-1, -1), "TOP"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#FFFFFF"), colors.HexColor("#FAFBFC")]),
        ("INNERGRID",    (0, 0), (-1, -1), 0.3, colors.HexColor("#DDDDDD")),
        ("BOX",          (0, 0), (-1, -1), 0.5, colors.HexColor("#DDDDDD")),
        ("TOPPADDING",   (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 4),
    ]
    for row_i, fill in enumerate(row_fills[1:], start=1):
        style_cmds.append(("BACKGROUND", (0, row_i), (-1, row_i), fill))

    tbl.setStyle(TableStyle(style_cmds))
    return tbl


def _pdf_match_bar(result: dict, width: int = 510):
    """Draw a horizontal match-outcome bar using ReportLab Drawing."""
    from reportlab.graphics.shapes import Drawing, Rect, String
    from reportlab.lib             import colors

    total   = result.get("total_source_rows", 0) or 1
    exact   = result.get("exact_matches", 0) or 0
    fuzzy   = (result.get("fuzzy_matches", 0) or 0) + (result.get("ai_matches", 0) or 0)
    exc_cnt = total - exact - fuzzy
    if exc_cnt < 0:
        exc_cnt = 0

    h    = 18
    d    = Drawing(width, h + 20)
    x    = 0
    segs = [
        (exact,   colors.HexColor("#2D7D46"), "Exact"),
        (fuzzy,   colors.HexColor("#b3ff00"), "Fuzzy/AI"),
        (exc_cnt, colors.HexColor("#C0392B"), "Exceptions"),
    ]
    label_y = 2
    for count, colour, label in segs:
        seg_w = (count / total) * width
        if seg_w < 1:
            x += seg_w
            continue
        d.add(Rect(x, label_y + 14, seg_w, h, fillColor=colour, strokeColor=None))
        if seg_w > 50:
            d.add(String(x + seg_w / 2, label_y + 20, f"{label}: {count}",
                         textAnchor="middle", fontSize=7,
                         fillColor=colors.HexColor("#1A1A1A")))
        x += seg_w
    return d


# ---------------------------------------------------------------------------
# Public: generate_single_pdf
# ---------------------------------------------------------------------------

def generate_single_pdf(result: dict) -> bytes:
    from reportlab.lib.pagesizes  import A4
    from reportlab.lib.units      import cm
    from reportlab.platypus       import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
    )
    from reportlab.lib            import colors

    result = copy.deepcopy(result)
    st     = _pdf_styles()
    cl     = st["colors"]
    buf    = io.BytesIO()

    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=1.8 * cm, rightMargin=1.8 * cm,
        topMargin=2 * cm, bottomMargin=2 * cm,
        title="Recon Agent — Reconciliation Report",
        author="Recon Agent",
    )

    story = []
    P     = Paragraph

    def hr():
        return HRFlowable(width="100%", thickness=0.5,
                          color=colors.HexColor("#DDDDDD"), spaceAfter=8, spaceBefore=4)

    def section(title: str):
        return P(title, st["section"])

    # ---- Branding + Title ----
    story.append(P("Recon Agent", st["muted"]))
    story.append(P("Reconciliation Report", st["title"]))
    story.append(P(
        f"<b>Source:</b> {result.get('source_filename','—')}  ·  "
        f"<b>Target:</b> {result.get('target_filename','—')}",
        st["subtitle"],
    ))
    story.append(P(
        f"Reconciled: {_fmt_date(result.get('completed_at'))}  ·  "
        f"Generated: {_now_str()}",
        st["muted"],
    ))
    story.append(hr())

    # ---- KPI Grid ----
    story.append(section("Key Metrics"))
    summary = result.get("summary") or {}
    kpis = [
        ("Total Transactions",  f"{result.get('total_source_rows', 0):,}"),
        ("Match Rate",          _fmt_pct(result.get("match_rate"))),
        ("Total Exceptions",    f"{result.get('exceptions_count', 0):,}"),
        ("Total Amount",        _fmt_inr(summary.get("total_amount"))),
        ("Matched Amount",      _fmt_inr(summary.get("matched_amount"))),
        ("Unmatched Amount",    _fmt_inr(summary.get("unmatched_amount"))),
    ]
    story.append(_pdf_kpi_table(kpis, col_count=3))
    story.append(Spacer(1, 12))

    # ---- Match breakdown ----
    story.append(section("Match Outcome"))
    story.append(_pdf_match_bar(result))
    story.append(Spacer(1, 6))
    story.append(P(
        f"Exact: <b>{result.get('exact_matches',0)}</b>  ·  "
        f"Fuzzy/AI: <b>{(result.get('fuzzy_matches',0) or 0) + (result.get('ai_matches',0) or 0)}</b>  ·  "
        f"Exceptions: <b>{result.get('exceptions_count',0)}</b>  ·  "
        f"AI Provider: <b>{result.get('ai_provider','none')}</b>",
        st["body"],
    ))
    story.append(P(
        f"Matching rules — Amount tolerance: <b>{_fmt_inr(result.get('amount_tolerance'))}</b>  ·  "
        f"Date window: <b>{result.get('date_window_days', 0)} days</b>",
        st["muted"],
    ))
    story.append(hr())

    # ---- Exception Analysis ----
    story.append(section("Exception Analysis"))

    merchants = summary.get("top_exception_merchants") or []
    if merchants:
        merch_data = [["Merchant / Party", "Exception Count"]]
        for m in merchants:
            merch_data.append([str(m.get("party", "")), str(m.get("count", 0))])
        merch_tbl = Table(merch_data, colWidths=[360, 150])
        merch_tbl.setStyle(TableStyle([
            ("BACKGROUND",   (0, 0), (-1, 0),  cl["dark"]),
            ("TEXTCOLOR",    (0, 0), (-1, 0),  cl["white"]),
            ("FONTNAME",     (0, 0), (-1, 0),  "Helvetica-Bold"),
            ("FONTSIZE",     (0, 0), (-1, -1), 8),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1),
             [colors.HexColor("#FFFFFF"), colors.HexColor("#FAFBFC")]),
            ("INNERGRID",    (0, 0), (-1, -1), 0.3, colors.HexColor("#DDDDDD")),
            ("BOX",          (0, 0), (-1, -1), 0.5, colors.HexColor("#DDDDDD")),
            ("TOPPADDING",   (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING",(0, 0), (-1, -1), 5),
        ]))
        story.append(merch_tbl)
        story.append(Spacer(1, 10))

    # Exceptions by date (table form for PDF reliability)
    exc_by_date = summary.get("exceptions_by_date") or []
    if len(exc_by_date) > 1:
        story.append(P("Exceptions by Date", st["section"]))
        date_data = [["Date", "Exception Count"]]
        for d in exc_by_date:
            date_data.append([
                str(d.get("date", "")).split("T")[0].split(" ")[0],
                str(d.get("count", 0)),
            ])
        date_tbl = Table(date_data, colWidths=[260, 250])
        date_tbl.setStyle(TableStyle([
            ("BACKGROUND",   (0, 0), (-1, 0),  cl["dark"]),
            ("TEXTCOLOR",    (0, 0), (-1, 0),  cl["white"]),
            ("FONTNAME",     (0, 0), (-1, 0),  "Helvetica-Bold"),
            ("FONTSIZE",     (0, 0), (-1, -1), 8),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1),
             [colors.HexColor("#FFFFFF"), colors.HexColor("#FAFBFC")]),
            ("INNERGRID",    (0, 0), (-1, -1), 0.3, colors.HexColor("#DDDDDD")),
            ("BOX",          (0, 0), (-1, -1), 0.5, colors.HexColor("#DDDDDD")),
            ("TOPPADDING",   (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING",(0, 0), (-1, -1), 5),
        ]))
        story.append(date_tbl)
        story.append(Spacer(1, 10))

    story.append(hr())

    # ---- Duplicate Summary ----
    dups = result.get("duplicates") or {}
    src_cnt = dups.get("source_count", 0) or 0
    tgt_cnt = dups.get("target_count", 0) or 0
    if src_cnt > 0 or tgt_cnt > 0:
        story.append(section("Potential Duplicate Detection"))
        story.append(P(
            f"Source file: <b>{src_cnt} duplicate group(s)</b>  ·  "
            f"Target file: <b>{tgt_cnt} duplicate group(s)</b>",
            st["body"],
        ))
        all_dups = []
        for d in (dups.get("source") or []):
            all_dups.append(["Source", d.get("party",""), _fmt_inr(d.get("amount")),
                              str(d.get("date","")), str(d.get("occurrences",""))])
        for d in (dups.get("target") or []):
            all_dups.append(["Target", d.get("party",""), _fmt_inr(d.get("amount")),
                              str(d.get("date","")), str(d.get("occurrences",""))])
        if all_dups:
            dup_header = [["File", "Party", "Amount", "Date", "Count"]]
            dup_tbl = Table(dup_header + all_dups, colWidths=[55, 160, 85, 75, 60], repeatRows=1)
            dup_tbl.setStyle(TableStyle([
                ("BACKGROUND",   (0, 0), (-1, 0),  cl["dark"]),
                ("TEXTCOLOR",    (0, 0), (-1, 0),  cl["white"]),
                ("FONTNAME",     (0, 0), (-1, 0),  "Helvetica-Bold"),
                ("FONTSIZE",     (0, 0), (-1, -1), 8),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1),
                 [colors.HexColor("#FDEBD0"), colors.HexColor("#FEF9E7")]),
                ("INNERGRID",    (0, 0), (-1, -1), 0.3, colors.HexColor("#DDDDDD")),
                ("BOX",          (0, 0), (-1, -1), 0.5, colors.HexColor("#DDDDDD")),
                ("TOPPADDING",   (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING",(0, 0), (-1, -1), 5),
            ]))
            story.append(dup_tbl)
        story.append(hr())

    # ---- Exception Details ----
    exceptions = result.get("exception_report") or []
    if exceptions:
        story.append(section(f"Exception Details ({len(exceptions)} records)"))
        if len(exceptions) > 200:
            story.append(P(
                f"Showing first 200 of {len(exceptions)} exceptions. "
                "Download the Excel report for the full list.",
                st["muted"],
            ))
        story.append(_pdf_exc_table(exceptions, st))

    doc.build(story)
    buf.seek(0)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Public: generate_batch_pdf
# ---------------------------------------------------------------------------

def generate_batch_pdf(batch: dict) -> bytes:
    from reportlab.lib.pagesizes  import A4
    from reportlab.lib.units      import cm
    from reportlab.platypus       import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
        HRFlowable, PageBreak,
    )
    from reportlab.lib            import colors

    batch = copy.deepcopy(batch)
    st    = _pdf_styles()
    cl    = st["colors"]
    buf   = io.BytesIO()

    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=1.8 * cm, rightMargin=1.8 * cm,
        topMargin=2 * cm, bottomMargin=2 * cm,
        title="Recon Agent — Batch Reconciliation Report",
        author="Recon Agent",
    )

    story = []
    P     = Paragraph

    def hr():
        return HRFlowable(width="100%", thickness=0.5,
                          color=colors.HexColor("#DDDDDD"), spaceAfter=8, spaceBefore=4)

    def section(title: str):
        return P(title, st["section"])

    summary = batch.get("summary") or {}
    runs    = batch.get("runs") or []

    # ---- Title ----
    story.append(P("Recon Agent", st["muted"]))
    story.append(P("Batch Reconciliation Report", st["title"]))
    story.append(P(
        f"Generated: {_now_str()}  ·  "
        f"{summary.get('completed_runs', 0)} completed, "
        f"{summary.get('failed_runs', 0)} failed",
        st["subtitle"],
    ))
    story.append(hr())

    # ---- Overall KPIs ----
    story.append(section("Overall Summary"))
    kpis = [
        ("Total Transactions",  f"{summary.get('total_transactions', 0):,}"),
        ("Match Rate",          _fmt_pct(summary.get("overall_match_rate"))),
        ("Total Exceptions",    f"{summary.get('total_exceptions', 0):,}"),
        ("Total Amount",        _fmt_inr(summary.get("total_amount"))),
        ("Matched Amount",      _fmt_inr(summary.get("total_matched_amount"))),
        ("Unmatched Amount",    _fmt_inr(summary.get("total_unmatched_amount"))),
    ]
    story.append(_pdf_kpi_table(kpis, col_count=3))
    story.append(Spacer(1, 12))

    dup_note = (
        f"Potential duplicate groups — Source: "
        f"<b>{summary.get('duplicate_source_groups', 0)}</b>  ·  "
        f"Target: <b>{summary.get('duplicate_target_groups', 0)}</b>"
    )
    story.append(P(dup_note, st["muted"]))
    story.append(hr())

    # ---- Per-run summary table ----
    story.append(section("Reconciliation Runs"))
    run_header = [["#", "Source File", "Target File", "Status",
                   "Txns", "Matched", "Match Rate", "Exceptions"]]
    run_rows   = []
    for i, run in enumerate(runs, start=1):
        status = run.get("status", "failed")
        r_data = run.get("result") or {}
        if status == "completed":
            run_rows.append([
                str(i),
                str(run.get("source_filename", ""))[:22],
                str(run.get("target_filename", ""))[:22],
                "✓ Complete",
                str(r_data.get("total_source_rows", 0)),
                str(r_data.get("total_matched", 0)),
                _fmt_pct(r_data.get("match_rate")),
                str(r_data.get("exceptions_count", 0)),
            ])
        else:
            run_rows.append([
                str(i),
                str(run.get("source_filename", ""))[:22],
                str(run.get("target_filename", ""))[:22],
                "✗ FAILED",
                "—", "—", "—", "—",
            ])

    run_tbl = Table(run_header + run_rows, colWidths=[22, 110, 110, 62, 35, 42, 52, 52], repeatRows=1)
    run_style = TableStyle([
        ("BACKGROUND",   (0, 0), (-1, 0),  cl["dark"]),
        ("TEXTCOLOR",    (0, 0), (-1, 0),  cl["white"]),
        ("FONTNAME",     (0, 0), (-1, 0),  "Helvetica-Bold"),
        ("FONTSIZE",     (0, 0), (-1, -1), 7),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [colors.HexColor("#FFFFFF"), colors.HexColor("#FAFBFC")]),
        ("INNERGRID",    (0, 0), (-1, -1), 0.3, colors.HexColor("#DDDDDD")),
        ("BOX",          (0, 0), (-1, -1), 0.5, colors.HexColor("#DDDDDD")),
        ("TOPPADDING",   (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 4),
        ("ALIGN",        (0, 0), (-1, -1), "LEFT"),
    ])
    # Highlight failed rows
    for i, run in enumerate(runs, start=1):
        if run.get("status") != "completed":
            run_style.add("BACKGROUND", (0, i), (-1, i), colors.HexColor("#FDEDEC"))
            run_style.add("TEXTCOLOR",  (0, i), (3, i), cl["red"])
    run_tbl.setStyle(run_style)
    story.append(run_tbl)

    # Failed run error details
    failed = [(i + 1, r) for i, r in enumerate(runs) if r.get("status") != "completed"]
    if failed:
        story.append(Spacer(1, 8))
        story.append(P("Failed Run Errors:", st["body"]))
        for idx, run in failed:
            story.append(P(
                f"<b>Run {idx}: {run.get('source_filename','')} ↔ "
                f"{run.get('target_filename','')}</b> — "
                f"{str(run.get('error','Unknown error'))[:200]}",
                st["muted"],
            ))

    story.append(hr())

    # ---- Per-run detail sections ----
    for i, run in enumerate(runs, start=1):
        if run.get("status") != "completed":
            continue
        r_data     = run.get("result") or {}
        exceptions = r_data.get("exception_report") or []
        if not exceptions:
            continue

        story.append(PageBreak())
        story.append(P(
            f"Run {i}: {run.get('source_filename', '')} ↔ {run.get('target_filename', '')}",
            st["title"],
        ))
        run_sum = r_data.get("summary") or {}
        story.append(P(
            f"Transactions: <b>{r_data.get('total_source_rows',0):,}</b>  ·  "
            f"Match rate: <b>{_fmt_pct(r_data.get('match_rate'))}</b>  ·  "
            f"Exceptions: <b>{r_data.get('exceptions_count',0)}</b>  ·  "
            f"Unmatched: <b>{_fmt_inr(run_sum.get('unmatched_amount'))}</b>",
            st["body"],
        ))
        story.append(Spacer(1, 8))
        story.append(_pdf_exc_table(exceptions, st))

    doc.build(story)
    buf.seek(0)
    return buf.getvalue()
