"""
tests/test_report_generator.py

Unit tests for core/report_generator.py.

Covers:
  - Single Excel: correct sheets, data, no crash
  - Single PDF: non-empty, valid PDF header
  - Batch Excel: correct sheet count
  - Batch PDF: non-empty
  - Empty exceptions: no crash
  - Failed batch pair: renders correctly
  - Currency formatting: ₹ symbol present in summary
  - Date formatting: ISO dates shortened correctly
  - Duplicate data: appears in both files
  - Input not mutated by report generation
"""

import copy
import pytest
from core.report_generator import (
    generate_single_excel,
    generate_single_pdf,
    generate_batch_excel,
    generate_batch_pdf,
)


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

SINGLE_RESULT = {
    "run_id":            "test-run-0001",
    "status":            "completed",
    "source_filename":   "source_jan.csv",
    "target_filename":   "target_jan.csv",
    "total_source_rows": 10,
    "total_matched":     8,
    "match_rate":        80.0,
    "exact_matches":     6,
    "fuzzy_matches":     2,
    "ai_matches":        0,
    "exceptions_count":  2,
    "exception_report":  [
        {
            "type":   "missing_target_record",
            "id":     "TXN-003",
            "party":  "Coffee Shop",
            "amount": 4.50,
            "reason": "No plausible match found",
            "date":   "2023-10-03",
        },
        {
            "type":   "stray_target_record",
            "id":     "REF-886",
            "party":  "Unknown Fee",
            "amount": 1.50,
            "reason": "Target record with no corresponding source record",
            "date":   "2023-10-06",
        },
    ],
    "ai_provider":      "none",
    "amount_tolerance": 20.0,
    "date_window_days": 5,
    "duplicates": {
        "source": [
            {
                "amount":      15.99,
                "party":       "Monthly Subscription",
                "date":        "2023-10-04",
                "occurrences": 2,
                "row_ids":     ["TXN-004", "TXN-004-DUP"],
            }
        ],
        "target":       [],
        "source_count": 1,
        "target_count": 0,
    },
    "summary": {
        "total_amount":            210.99,
        "matched_amount":          204.99,
        "unmatched_amount":        6.00,
        "top_exception_merchants": [
            {"party": "Coffee Shop", "count": 1},
        ],
        "exceptions_by_date": [
            {"date": "2023-10-03", "count": 1},
            {"date": "2023-10-06", "count": 1},
        ],
    },
    "completed_at": "2023-10-10T12:00:00+00:00",
}

BATCH_RESULT = {
    "summary": {
        "total_transactions":    10,
        "total_matched":         8,
        "total_exceptions":      2,
        "overall_match_rate":    80.0,
        "total_amount":          210.99,
        "total_matched_amount":  204.99,
        "total_unmatched_amount": 6.00,
        "duplicate_source_groups": 1,
        "duplicate_target_groups": 0,
        "completed_runs":        1,
        "failed_runs":           1,
    },
    "runs": [
        {
            "source_filename": "source_jan.csv",
            "target_filename": "target_jan.csv",
            "status":          "completed",
            "error":           None,
            "result":          SINGLE_RESULT,
        },
        {
            "source_filename": "source_feb.csv",
            "target_filename": "target_feb.csv",
            "status":          "failed",
            "error":           "Column 'amount' not found in source file.",
            "result":          None,
        },
    ],
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def parse_excel(data: bytes):
    from openpyxl import load_workbook
    import io
    return load_workbook(io.BytesIO(data))


# ---------------------------------------------------------------------------
# Single Excel
# ---------------------------------------------------------------------------

class TestSingleExcel:

    def test_returns_bytes(self):
        result = generate_single_excel(SINGLE_RESULT)
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_four_sheets_present(self):
        wb = parse_excel(generate_single_excel(SINGLE_RESULT))
        names = wb.sheetnames
        assert "Summary" in names
        assert "Exceptions" in names
        assert "Duplicates" in names
        assert "Top Exception Merchants" in names

    def test_summary_sheet_contains_source_filename(self):
        wb = parse_excel(generate_single_excel(SINGLE_RESULT))
        ws = wb["Summary"]
        all_values = [str(c.value) for row in ws.iter_rows() for c in row if c.value]
        assert any("source_jan.csv" in v for v in all_values)

    def test_summary_sheet_contains_match_rate(self):
        wb = parse_excel(generate_single_excel(SINGLE_RESULT))
        ws = wb["Summary"]
        # Match rate stored as 0.80 with percentage format
        all_values = [c.value for row in ws.iter_rows() for c in row if c.value is not None]
        assert 0.80 in all_values or any(v == 0.80 for v in all_values)

    def test_exceptions_sheet_has_correct_row_count(self):
        wb = parse_excel(generate_single_excel(SINGLE_RESULT))
        ws = wb["Exceptions"]
        # 1 header + 2 exceptions
        assert ws.max_row == 3

    def test_exceptions_sheet_party_present(self):
        wb = parse_excel(generate_single_excel(SINGLE_RESULT))
        ws = wb["Exceptions"]
        all_values = [str(c.value) for row in ws.iter_rows() for c in row if c.value]
        assert any("Coffee Shop" in v for v in all_values)

    def test_duplicates_sheet_has_source_group(self):
        wb = parse_excel(generate_single_excel(SINGLE_RESULT))
        ws = wb["Duplicates"]
        all_values = [str(c.value) for row in ws.iter_rows() for c in row if c.value]
        assert any("Source" in v for v in all_values)
        assert any("Monthly Subscription" in v for v in all_values)

    def test_top_merchants_sheet_has_data(self):
        wb = parse_excel(generate_single_excel(SINGLE_RESULT))
        ws = wb["Top Exception Merchants"]
        all_values = [str(c.value) for row in ws.iter_rows() for c in row if c.value]
        assert any("Coffee Shop" in v for v in all_values)

    def test_does_not_mutate_input(self):
        original = copy.deepcopy(SINGLE_RESULT)
        generate_single_excel(SINGLE_RESULT)
        assert SINGLE_RESULT == original


# ---------------------------------------------------------------------------
# Single PDF
# ---------------------------------------------------------------------------

class TestSinglePDF:

    def test_returns_bytes(self):
        result = generate_single_pdf(SINGLE_RESULT)
        assert isinstance(result, bytes)
        assert len(result) > 100

    def test_valid_pdf_header(self):
        result = generate_single_pdf(SINGLE_RESULT)
        assert result[:4] == b"%PDF"

    def test_does_not_mutate_input(self):
        original = copy.deepcopy(SINGLE_RESULT)
        generate_single_pdf(SINGLE_RESULT)
        assert SINGLE_RESULT == original


# ---------------------------------------------------------------------------
# Batch Excel
# ---------------------------------------------------------------------------

class TestBatchExcel:

    def test_returns_bytes(self):
        result = generate_batch_excel(BATCH_RESULT)
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_contains_batch_overview_sheet(self):
        wb = parse_excel(generate_batch_excel(BATCH_RESULT))
        assert "Batch Overview" in wb.sheetnames

    def test_contains_run_summary_sheet(self):
        wb = parse_excel(generate_batch_excel(BATCH_RESULT))
        assert "Run Summary" in wb.sheetnames

    def test_run_summary_has_correct_row_count(self):
        wb = parse_excel(generate_batch_excel(BATCH_RESULT))
        ws = wb["Run Summary"]
        # 1 header + 2 runs
        assert ws.max_row == 3

    def test_failed_run_shows_failed_status(self):
        wb = parse_excel(generate_batch_excel(BATCH_RESULT))
        ws = wb["Run Summary"]
        all_values = [str(c.value) for row in ws.iter_rows() for c in row if c.value]
        assert any("FAILED" in v for v in all_values)

    def test_per_run_exception_sheet_created_for_completed_runs(self):
        wb = parse_excel(generate_batch_excel(BATCH_RESULT))
        # The completed run has exceptions, so a detail sheet should exist
        names = wb.sheetnames
        per_run_sheets = [n for n in names if n.startswith("R0")]
        assert len(per_run_sheets) == 1

    def test_does_not_mutate_input(self):
        original = copy.deepcopy(BATCH_RESULT)
        generate_batch_excel(BATCH_RESULT)
        assert BATCH_RESULT == original


# ---------------------------------------------------------------------------
# Batch PDF
# ---------------------------------------------------------------------------

class TestBatchPDF:

    def test_returns_bytes(self):
        result = generate_batch_pdf(BATCH_RESULT)
        assert isinstance(result, bytes)
        assert len(result) > 100

    def test_valid_pdf_header(self):
        result = generate_batch_pdf(BATCH_RESULT)
        assert result[:4] == b"%PDF"

    def test_does_not_mutate_input(self):
        original = copy.deepcopy(BATCH_RESULT)
        generate_batch_pdf(BATCH_RESULT)
        assert BATCH_RESULT == original


# ---------------------------------------------------------------------------
# Edge cases
# ---------------------------------------------------------------------------

class TestEdgeCases:

    def test_empty_exceptions_no_crash_excel(self):
        result = copy.deepcopy(SINGLE_RESULT)
        result["exception_report"] = []
        result["exceptions_count"] = 0
        data = generate_single_excel(result)
        wb = parse_excel(data)
        ws = wb["Exceptions"]
        # Only header row
        assert ws.max_row == 1

    def test_empty_exceptions_no_crash_pdf(self):
        result = copy.deepcopy(SINGLE_RESULT)
        result["exception_report"] = []
        result["exceptions_count"] = 0
        data = generate_single_pdf(result)
        assert data[:4] == b"%PDF"

    def test_no_duplicates_no_crash(self):
        result = copy.deepcopy(SINGLE_RESULT)
        result["duplicates"] = {
            "source": [], "target": [],
            "source_count": 0, "target_count": 0,
        }
        data = generate_single_excel(result)
        wb = parse_excel(data)
        ws = wb["Duplicates"]
        assert ws.max_row >= 1  # at least header or empty message

    def test_batch_all_failed_no_crash(self):
        batch = copy.deepcopy(BATCH_RESULT)
        for run in batch["runs"]:
            run["status"] = "failed"
            run["result"] = None
        data = generate_batch_excel(batch)
        assert len(data) > 0
        data_pdf = generate_batch_pdf(batch)
        assert data_pdf[:4] == b"%PDF"

    def test_missing_completed_at_no_crash(self):
        result = copy.deepcopy(SINGLE_RESULT)
        result["completed_at"] = None
        data = generate_single_pdf(result)
        assert data[:4] == b"%PDF"

    def test_inr_formatting(self):
        from core.report_generator import _fmt_inr
        assert _fmt_inr(1234567.89) == "₹1,234,567.89"
        assert _fmt_inr(0.0) == "₹0.00"
        assert _fmt_inr(None) == "—"
        assert _fmt_inr(-500.0).startswith("−₹")

    def test_date_formatting(self):
        from core.report_generator import _fmt_date
        result = _fmt_date("2023-10-10T12:00:00+00:00")
        assert "2023" in result
        assert _fmt_date(None) == ""
        assert _fmt_date("") == ""
